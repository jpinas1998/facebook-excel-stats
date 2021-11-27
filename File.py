import locale
import operator
import os

from tkinter import Tk
from tkinter.filedialog import askopenfilename, asksaveasfilename

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from Post import Post
from Config import Config


def find_row_in_talking(talking_about_sheet, id):
    for col in talking_about_sheet.iter_cols(min_col=2, max_col=2, min_row=2):
        for cell in col:
            if cell.value == id:
                return cell.row


def obtener_posts(workbook):
    conf = Config()

    # obtengo la hoja activa, en este caso es siempre la primera
    key_metrics_sheet = workbook.get_sheet_by_name("Key Metrics")
    talking_about_sheet = workbook.get_sheet_by_name("Lifetime Talking About This(...")

    last_row = key_metrics_sheet.max_row
    last_colum = 18  # sheet.max_column

    # xq los 2 primeras filas son bla bla bla
    posts = []
    for row in key_metrics_sheet.iter_rows(min_row=3, max_col=last_colum, max_row=last_row):
        id = row[conf.ID].value
        enlace = row[conf.ENLACE].value
        mensaje = row[conf.MENSAJE].value
        tipo = row[conf.TIPO].value
        fecha = row[conf.FECHA_PUBLICACION].value
        alcance = row[conf.ALCANCE].value
        engagement = row[conf.ENGAGEMENT].value
        match_audience = row[conf.MATCH_AUDIENCE].value
        feed_negativo = row[conf.FEED_BACK_NEGATIVE].value

        # busco la fila del post para obtener los comentarios
        matching_row = find_row_in_talking(talking_about_sheet, id)

        # obtener las métricas de la pag encontrada
        comments = talking_about_sheet.cell(row=matching_row, column=conf.COMMENT).value
        likes = talking_about_sheet.cell(row=matching_row, column=conf.LIKE).value
        shared = talking_about_sheet.cell(row=matching_row, column=conf.SHARED).value

        posts.append(Post(
            id,
            enlace,
            mensaje,
            tipo,
            fecha,
            alcance,
            comments,
            shared,
            likes,
            engagement,
            match_audience,
            feed_negativo,
        ))
    return posts


def obtener_top(post, reverse=True, tipo="personas_alcanzadas", top=3):
    # lista resultante
    result_list = []

    # lista ordenada por los campos de insterés
    order_list = sorted(post, key=operator.attrgetter(tipo), reverse=reverse)

    # pueblo la lista resultante con el top de los elementos
    for i in range(0, top):
        result_list.append(order_list[i])
    return result_list


def save(posts, base_path, top=3):
    conf = Config()

    posts_ordenados_por_fecha = obtener_top(posts, tipo="fecha", top=len(posts))
    end_date = posts_ordenados_por_fecha[0].fecha
    start_date = posts_ordenados_por_fecha[-1].fecha

    wb = Workbook()
    locale.setlocale(locale.LC_TIME, '')
    dest_filename = 'stats SJ del ' + start_date.strftime('%d %b') + ' al ' + end_date.strftime('%d %b') + '.xlsx'

    # pido la fecha para poner en el mensaje de título

    title = "Estadísticas de SJ en el período {0}/{1}".format(
        start_date.strftime('%Y-%m-%d'),
        end_date.strftime('%Y-%m-%d')
    )

    ws_mas_completos = wb.active
    ws_mas_completos.title = "Mejores posts"
    escribir_posts(
        ws_mas_completos,
        title,
        posts,
        "indice_completo",
        top,
        "Post más completos según el índice (intereacciones*{0} + match audience*{1} + engagement*{2} + "
        "alcance*{3})".format(conf.INTERACCIONES_WEIGHT, conf.MATCH_AUDIENCE_WEIGHT, conf.ENGAGEMENT_WEIGHT,
                              conf.ALCANCE_WEIGHT),
        "Post menos completos según el índice (intereacciones*{0} + match audience*{1} + engagement*{2} + "
        "alcance*{3})".format(conf.INTERACCIONES_WEIGHT, conf.MATCH_AUDIENCE_WEIGHT, conf.ENGAGEMENT_WEIGHT,
                              conf.ALCANCE_WEIGHT)
    )

    ws_interacciones = wb.create_sheet("Interacciones")
    conf.show_alcance = False
    conf.show_comments = False
    conf.show_reacciones = False
    conf.show_compartidos = False
    conf.show_engagement = False
    conf.show_match_audience = False
    conf.show_indice_completo = False
    conf.show_feed_back_negativo = False
    escribir_posts(
        ws_interacciones,
        title,
        posts,
        "indice_interacciones",
        top,
        "Post con más interacciones (reacciones*{0} + comentarios*{1} + compartido*{2}".format(
            conf.LIKE_WEIGHT,
            conf.COMMENT_WEIGHT,
            conf.SHARED_WEIGHT
        ),
        "Post con más interacciones (reacciones*{0} + comentarios*{1} + compartido*{2}".format(
            conf.LIKE_WEIGHT,
            conf.COMMENT_WEIGHT,
            conf.SHARED_WEIGHT
        )
    )

    conf.show_alcance = True
    conf.show_comments = False
    conf.show_reacciones = False
    conf.show_compartidos = False
    conf.show_indice_interacciones = False
    conf.show_engagement = False
    conf.show_match_audience = False
    conf.show_indice_completo = False
    conf.show_feed_back_negativo = False

    ws_alcance = wb.create_sheet("Alcance")
    escribir_posts(
        ws_alcance,
        title,
        posts,
        "personas_alcanzadas",
        top,
        "Post con más alcance",
        "Post con menos alcance"
    )

    conf.show_alcance = False
    conf.show_comments = True
    conf.show_reacciones = False
    conf.show_compartidos = False
    conf.show_indice_interacciones = False
    conf.show_engagement = False
    conf.show_match_audience = False
    conf.show_indice_completo = False
    conf.show_feed_back_negativo = False

    ws_comments = wb.create_sheet("Comentarios")
    escribir_posts(
        ws_comments,
        title,
        posts,
        "cant_coment",
        top,
        "Post con más comentarios",
        "Post con menos comentrios"
    )

    conf.show_alcance = False
    conf.show_comments = False
    conf.show_reacciones = True
    conf.show_compartidos = False
    conf.show_indice_interacciones = False
    conf.show_engagement = False
    conf.show_match_audience = False
    conf.show_indice_completo = False
    conf.show_feed_back_negativo = False

    ws_likes = wb.create_sheet("Reacciones")
    escribir_posts(
        ws_likes,
        title,
        posts,
        "cant_likes",
        top,
        "Post con más reacciones",
        "Post con menos reacciones"
    )

    conf.show_alcance = False
    conf.show_comments = False
    conf.show_reacciones = False
    conf.show_compartidos = True
    conf.show_indice_interacciones = False
    conf.show_engagement = False
    conf.show_match_audience = False
    conf.show_indice_completo = False
    conf.show_feed_back_negativo = False

    ws_compartido = wb.create_sheet("Compartido")
    escribir_posts(
        ws_compartido,
        title,
        posts,
        "cant_share",
        top,
        "Post con más compartidos",
        "Post con menos compartidos"
    )

    conf.show_alcance = True
    conf.show_comments = True
    conf.show_reacciones = True
    conf.show_compartidos = True
    conf.show_indice_interacciones = True
    conf.show_engagement = True
    conf.show_match_audience = True
    conf.show_indice_completo = True
    conf.show_feed_back_negativo = True

    #  salvando todos los post
    ws_all_post = wb.create_sheet("All Posts")
    write_posts(ws_all_post, posts)

    final_path = os.path.join(base_path, dest_filename)
    if os.path.exists(final_path):
        os.remove(final_path)

    wb.save(filename=final_path)


def escribir_posts(ws, titulo, posts, tipo, top, mensaje_best, mensaje_worse):
    # POSTS TOP
    top_best_post = obtener_top(posts, tipo=tipo, top=top)
    top_worse_post = obtener_top(posts, reverse=False, tipo=tipo, top=top)

    # escribo el título de la hoja
    ws.cell(
        row=1,
        column=1,
        value=titulo
    ).font = Font(bold=True, size=15)

    # salvando según indice más completo
    write_posts(
        ws,
        top_best_post,
        base_col=1,
        mensaje=mensaje_best
    )
    write_posts(
        ws,
        top_worse_post,
        base_col=1,
        mensaje=mensaje_worse
    )


def write_table_header(ws, base_row, base_col, mensaje):
    conf = Config()
    if len(mensaje) != 0:
        ws.cell(column=base_col, row=base_row, value=mensaje).font = Font(bold=True, size="10")
        base_row += 1

    ws.cell(column=base_col, row=base_row, value="Id").font = Font(bold=True)
    ws.column_dimensions[get_column_letter(base_col)].width = 3
    base_col += 1
    ws.cell(column=base_col, row=base_row, value="Enlace").font = Font(bold=True)
    ws.column_dimensions[get_column_letter(base_col)].width = 7
    base_col += 1
    ws.cell(column=base_col, row=base_row, value="Tipo").font = Font(bold=True)
    ws.column_dimensions[get_column_letter(base_col)].width = 11
    base_col += 1
    ws.cell(column=base_col, row=base_row, value="Mensaje").font = Font(bold=True)
    ws.column_dimensions[get_column_letter(base_col)].width = 20
    base_col += 1
    ws.cell(column=base_col, row=base_row, value="Palabras").font = Font(bold=True)
    base_col += 1
    ws.cell(column=base_col, row=base_row, value="Emoji").font = Font(bold=True)
    base_col += 1
    ws.cell(column=base_col, row=base_row, value="Hashtag").font = Font(bold=True)
    base_col += 1
    ws.cell(column=base_col, row=base_row, value="Fecha").font = Font(bold=True)
    ws.column_dimensions[get_column_letter(base_col)].width = 12
    base_col += 1
    ws.cell(column=base_col, row=base_row, value="Día de la semana").font = Font(bold=True)
    ws.column_dimensions[get_column_letter(base_col)].width = len(str(ws.cell(column=base_col, row=base_row).value)) - 3
    base_col += 1
    ws.cell(column=base_col, row=base_row, value="Horario").font = Font(bold=True)
    ws.column_dimensions[get_column_letter(base_col)].width = len(str(ws.cell(column=base_col, row=base_row).value))
    base_col += 1

    if conf.show_alcance:
        ws.cell(column=base_col, row=base_row, value="Alcance").font = Font(bold=True)
        ws.column_dimensions[get_column_letter(base_col)].width = len(str(ws.cell(column=base_col, row=base_row).value))
        base_col += 1
    if conf.show_comments:
        ws.cell(column=base_col, row=base_row, value="Comentarios").font = Font(bold=True)
        ws.column_dimensions[get_column_letter(base_col)].width = len(str(ws.cell(column=base_col, row=base_row).value))
        base_col += 1
    if conf.show_reacciones:
        ws.cell(column=base_col, row=base_row, value="Reacciones").font = Font(bold=True)
        ws.column_dimensions[get_column_letter(base_col)].width = len(str(ws.cell(column=base_col, row=base_row).value))
        base_col += 1
    if conf.show_compartidos:
        ws.cell(column=base_col, row=base_row, value="Compartidos").font = Font(bold=True)
        ws.column_dimensions[get_column_letter(base_col)].width = len(str(ws.cell(column=base_col, row=base_row).value))
        base_col += 1
    if conf.show_indice_interacciones:
        ws.cell(column=base_col, row=base_row, value="Índice interacciones").font = Font(bold=True)
        ws.column_dimensions[get_column_letter(base_col)].width = len(str(ws.cell(column=base_col, row=base_row).value)) - 3
        base_col += 1
    if conf.show_engagement:
        ws.cell(column=base_col, row=base_row, value="Engagement").font = Font(bold=True)
        ws.column_dimensions[get_column_letter(base_col)].width = len(str(ws.cell(column=base_col, row=base_row).value))
        base_col += 1
    if conf.show_match_audience:
        ws.cell(column=base_col, row=base_row, value="Match audience").font = Font(bold=True)
        ws.column_dimensions[get_column_letter(base_col)].width = len(str(ws.cell(column=base_col, row=base_row).value))
        base_col += 1
    if conf.show_indice_completo:
        ws.cell(column=base_col, row=base_row, value="Índice completo").font = Font(bold=True)
        ws.column_dimensions[get_column_letter(base_col)].width = len(str(ws.cell(column=base_col, row=base_row).value))
        base_col += 1
    if conf.show_feed_back_negativo:
        ws.cell(column=base_col, row=base_row, value="Feed back negativo").font = Font(bold=True)
        ws.column_dimensions[get_column_letter(base_col)].width = len(str(ws.cell(column=base_col, row=base_row).value))

    return base_row


def write_posts(ws, posts, mensaje="", base_col=1):
    conf = Config()
    latest_row = ws.max_row
    if ws.cell(row=1, column=1).value:  # para crear un espacio entre las tablas
        latest_row += 2

    latest_row = write_table_header(ws, latest_row, base_col, mensaje=mensaje) + 1

    for row, post in zip(range(latest_row, latest_row + len(posts)), posts):
        col = base_col
        ws.cell(column=col, row=row, value=post.id)
        col += 1
        ws.cell(column=col, row=row, value=post.enlace)
        col += 1
        ws.cell(column=col, row=row, value=post.tipo)
        col += 1
        ws.cell(column=col, row=row, value=post.mensaje)
        col += 1
        ws.cell(column=col, row=row, value=post.count_words())
        col += 1
        ws.cell(column=col, row=row, value=post.text_has_emoji())
        col += 1
        ws.cell(column=col, row=row, value=post.text_has_hashtag())
        col += 1
        ws.cell(column=col, row=row, value=post.get_dia_mes_ano())
        col += 1
        ws.cell(column=col, row=row, value=post.get_week_day())
        col += 1
        ws.cell(column=col, row=row, value=post.find_horario())
        col += 1

        if conf.show_alcance:
            ws.cell(column=col, row=row, value=post.personas_alcanzadas)
            col += 1
        if conf.show_comments:
            ws.cell(column=col, row=row, value=post.cant_coment)
            col += 1
        if conf.show_reacciones:
            ws.cell(column=col, row=row, value=post.cant_likes)
            col += 1
        if conf.show_compartidos:
            ws.cell(column=col, row=row, value=post.cant_share)
            col += 1
        if conf.show_indice_interacciones:
            ws.cell(column=col, row=row, value=post.indice_interacciones)
            col += 1
        if conf.show_engagement:
            ws.cell(column=col, row=row, value=post.engagement)
            col += 1
        if conf.show_match_audience:
            ws.cell(column=col, row=row, value=post.matching_audience)
            col += 1
        if conf.show_indice_completo:
            ws.cell(column=col, row=row, value=post.indice_completo)
            col += 1
        if conf.show_feed_back_negativo:
            ws.cell(column=col, row=row, value=post.feed_back_negativo)


def open_file_explorer():
    mask = [
        ("Excel files", "*.xlsx"),
    ]
    Tk().withdraw()  # we don't want a full GUI, so keep the root window from appearin
    return askopenfilename(filetypes=mask)


def export_file():
    mask = [
        ("Excel files", "*.xlsx"),
    ]
    Tk().withdraw()  # we don't want a full GUI, so keep the root window from appearin
    filename = asksaveasfilename(filetypes=mask)
    if filename:
        if ".xlsx" not in filename:
            filename += ".xlsx"
    return filename


def is_valid_excel(path):
    is_facebook_excel = True
    try:
        excel = load_workbook(path)
        excel.get_sheet_by_name("Key Metrics")
        excel.get_sheet_by_name("Lifetime Talking About This(...")
    except:
        is_facebook_excel = False
    return is_facebook_excel
